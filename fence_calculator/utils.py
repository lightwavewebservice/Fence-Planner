from __future__ import annotations
from decimal import Decimal, ROUND_UP
from typing import Dict, Any, Optional
from django.conf import settings
from .models import FenceType, Material, FenceCalculation
from django.utils import timezone

# -------- Calculation Logic -------- #

def quantize_2(d: Decimal) -> Decimal:
    return d.quantize(Decimal('0.01'))


def ceil_div(a: Decimal, b: Decimal) -> int:
    # Ceiling for Decimal division
    return int((a / b).to_integral_value(rounding=ROUND_UP))


def calculate_fence_requirements(
    fence_type: FenceType,
    fence_length: Decimal,
    labor_rate: Decimal | None = None,
    build_rate: Decimal | None = None,
    price_overrides: Dict[str, Decimal] | None = None,
    *,
    top_wire_type: str | None = None,  # 'standard' | 'hot' | 'barb'
    insulator_type: str | None = None,  # 'claw' | 'bullnose' (only for hot)
    post_spacing_override: Decimal | None = None,
    terrain: str | None = None,  # 'flat' | 'broken' (undulating/broken)
    wire_count_override: int | None = None,
    hot_wire_count: int | None = None,  # number of hot wires when top wire is 'hot'
    staples_per_box: int | None = None,
    netting_type: str = 'none',
    electric_outrigger: bool = False,
) -> Dict[str, Any]:
    price_overrides = price_overrides or {}

    # Quantities
    spacing = post_spacing_override if post_spacing_override else fence_type.post_spacing
    posts_required = ceil_div(fence_length, spacing) + 1

    netting_type = (netting_type or 'none').lower()
    if netting_type not in {'none', 'sheep', 'deer'}:
        netting_type = 'none'

    netting_height_cm: Optional[Decimal]
    if netting_type == 'deer':
        netting_height_cm = Decimal('200')
    elif netting_type == 'sheep':
        netting_height_cm = Decimal('120')
    else:
        netting_height_cm = None

    # Normalize selection
    twt = (top_wire_type or 'standard').lower()
    if twt not in ('standard', 'hot', 'barb'):
        twt = 'standard'

    # Compute wire lengths/rolls considering the top wire selection
    if wire_count_override is not None:
        total_wires = max(int(wire_count_override), 0)
    else:
        total_wires = int(fence_type.wire_count or 0)
    # Determine special wires based on type and requested hot wire count
    if twt == 'hot' and hot_wire_count and total_wires > 0:
        special_wires = min(int(hot_wire_count), total_wires)
    elif twt == 'barb' and total_wires > 0:
        special_wires = 1
    elif twt == 'hot' and total_wires > 0:
        # Default to 1 hot wire if not specified
        special_wires = 1
    else:
        special_wires = 0
    standard_wires = max(total_wires - special_wires, 0)

    standard_wire_len_m = quantize_2(Decimal(standard_wires) * fence_length)
    special_wire_len_m = quantize_2(Decimal(special_wires) * fence_length)

    # Use material-specific roll length or fall back to setting
    def get_roll_length(material):
        if material and material.roll_length:
            return Decimal(str(material.roll_length))
        return Decimal(str(settings.WIRE_ROLL_LENGTH))

    standard_wire_rolls = Decimal(0)
    if standard_wire_len_m > 0 and fence_type.wire_material:
        roll_len = get_roll_length(fence_type.wire_material)
        standard_wire_rolls = Decimal(ceil_div(standard_wire_len_m, roll_len))

    special_wire_rolls = Decimal(0)
    special_wire_material = None
    if special_wire_len_m > 0:
        # choose material based on top wire type
        if twt == 'hot':
            # For hot wire, use the same material as standard wires
            special_wire_material = fence_type.wire_material
        elif twt == 'barb':
            special_wire_material = fence_type.barb_wire_material or fence_type.wire_material
        
        if special_wire_material:
            roll_len = get_roll_length(special_wire_material)
            special_wire_rolls = Decimal(ceil_div(special_wire_len_m, roll_len))

    wire_length_meters = quantize_2(Decimal(total_wires) * fence_length)
    # Note: wire_rolls_required will be calculated after combining logic

    # Labor
    lr = Decimal(str(labor_rate)) if labor_rate is not None else Decimal(str(settings.LABOR_RATE_PER_HOUR))
    br = Decimal(str(build_rate)) if build_rate is not None else Decimal(str(settings.BUILD_RATE_METERS_PER_HOUR))
    labor_hours = quantize_2(fence_length / br)
    labor_cost = quantize_2(labor_hours * lr)

    # Material prices helper
    def eff_price(material: Material | None) -> Decimal:
        if not material:
            return Decimal('0.00')
        override_key = str(material.id)
        if override_key in price_overrides:
            return Decimal(str(price_overrides[override_key]))
        return Decimal(material.current_price or material.default_price)

    material_costs: Dict[str, Any] = {}

    # Posts cost (per unit)
    if fence_type.post_material:
        p = eff_price(fence_type.post_material)
        material_costs['posts'] = {
            'material': fence_type.post_material.name,
            'unit_price': float(quantize_2(p)),
            'quantity': posts_required,
            'cost': float(quantize_2(p * Decimal(posts_required))),
        }

    # Wire cost (assume price per roll). Combine standard and special wire if same material
    total_standard_wire_rolls = standard_wire_rolls
    
    # If special wire uses same material as standard, combine them
    if (special_wire_material and fence_type.wire_material and 
        special_wire_material.id == fence_type.wire_material.id):
        # Combine standard and special wire rolls
        total_standard_wire_rolls = standard_wire_rolls + special_wire_rolls
        special_wire_rolls = Decimal(0)  # Reset special rolls since we combined them
        special_wire_material = None  # Clear special material

    # Add standard wire (possibly combined with special)
    if fence_type.wire_material and total_standard_wire_rolls > 0:
        p = eff_price(fence_type.wire_material)
        material_costs['wire_standard'] = {
            'material': fence_type.wire_material.name,
            'unit_price': float(quantize_2(p)),
            'quantity': float(total_standard_wire_rolls),
            'cost': float(quantize_2(p * total_standard_wire_rolls)),
        }

    # Add special wire only if it's different from standard wire
    if special_wire_rolls > 0 and special_wire_material:
        p = eff_price(special_wire_material)
        material_costs['wire_top_' + twt] = {
            'material': special_wire_material.name,
            'unit_price': float(quantize_2(p)),
            'quantity': float(special_wire_rolls),
            'cost': float(quantize_2(p * special_wire_rolls)),
        }
    
    # Calculate total wire rolls required (after combining logic)
    wire_rolls_required = total_standard_wire_rolls + special_wire_rolls


    # Netting (if configured) — calculate based on fence length and roll length
    netting_rolls_required: Decimal = Decimal('0')
    if fence_type.netting_material:
        p = eff_price(fence_type.netting_material)
        netting_roll_length = fence_type.netting_material.roll_length or Decimal('50')
        netting_roll_length = Decimal(str(netting_roll_length))
        netting_rolls_required = Decimal(ceil_div(fence_length, netting_roll_length))
        if netting_rolls_required > 0:
            material_costs['netting'] = {
                'material': fence_type.netting_material.name,
                'unit_price': float(quantize_2(p)),
                'quantity': float(netting_rolls_required),
                'cost': float(quantize_2(p * netting_rolls_required)),
            }

    # Staples (for non-hot wires and netting)
    staple_counts: Dict[str, int] = {}
    if getattr(settings, 'STAPLES_ENABLED', True):
        # Wires that get staples: all wires unless hot wires are present, then only standard wires
        stapled_wires = total_wires if twt != 'hot' else standard_wires
        # Posts distribution
        end_posts = 2 if posts_required >= 2 else posts_required
        line_posts = max(posts_required - end_posts, 0)
        # Per-post usage
        per_line = int(getattr(settings, 'STAPLES_PER_WIRE_PER_LINE_POST', 1))
        per_end = int(getattr(settings, 'STAPLES_PER_WIRE_PER_END_POST', 2))
        per_net = int(getattr(settings, 'STAPLES_PER_POST_FOR_NETTING', 4))
        # Counts
        line_staples = line_posts * stapled_wires * per_line
        end_staples = end_posts * stapled_wires * per_end
        netting_staples = (posts_required * per_net) if fence_type.netting_material else 0
        total_staples = int(line_staples + end_staples + netting_staples)
        # Boxes
        spb = int(staples_per_box) if (staples_per_box and int(staples_per_box) > 0) else int(getattr(settings, 'STAPLES_PER_BOX', 2000))
        boxes = int((Decimal(total_staples) / Decimal(spb)).to_integral_value(rounding=ROUND_UP)) if spb > 0 else 0
        staple_counts = {
            'staples_per_box_used': spb,
            'line_staples': line_staples,
            'end_staples': end_staples,
            'netting_staples': netting_staples,
            'total_staples': total_staples,
            'boxes': boxes,
        }

        # Material cost for staples
        staples_mat = Material.objects.filter(name__iexact=getattr(settings, 'STAPLES_MATERIAL_NAME', 'U Staples'), is_active=True).first()
        if staples_mat:
            p = eff_price(staples_mat)
            staples_name = staples_mat.name
        else:
            p = Decimal(str(getattr(settings, 'STAPLES_DEFAULT_PRICE', 0)))
            staples_name = getattr(settings, 'STAPLES_MATERIAL_NAME', 'U Staples')
        if boxes > 0:
            material_costs['staples'] = {
                'material': staples_name,
                'unit_price': float(quantize_2(p)),
                'quantity': boxes,
                'cost': float(quantize_2(p * Decimal(boxes))),
            }

    # Insulators for hot wires
    insulator_counts: Dict[str, int] = {}
    if twt == 'hot':
        # Use number of hot wires determined above
        num_hot_wires = special_wires
        # Bullnose at each end per hot wire (2 per wire), claw on each intermediate post per hot wire
        bullnose_per_wire = 2 if posts_required >= 2 else max(posts_required, 0)
        claw_per_wire = max(posts_required - 2, 0)
        bullnose_qty = bullnose_per_wire * num_hot_wires
        claw_qty = claw_per_wire * num_hot_wires
        insulator_counts = {'bullnose': bullnose_qty, 'claw': claw_qty, 'hot_wires': num_hot_wires}

        bullnose_mat = Material.objects.filter(name='Bullnose Insulator', is_active=True).first()
        if bullnose_mat and bullnose_qty > 0:
            p = eff_price(bullnose_mat)
            # Optional: label to reflect multiple hot wires
            material_label = bullnose_mat.name
            if num_hot_wires > 1:
                material_label = f"{bullnose_mat.name} (for {num_hot_wires} hot wires)"
            elif num_hot_wires == 1:
                material_label = f"{bullnose_mat.name} (for hot wire)"
            material_costs['insulators_bullnose'] = {
                'material': material_label,
                'unit_price': float(quantize_2(p)),
                'quantity': bullnose_qty,
                'cost': float(quantize_2(p * Decimal(bullnose_qty))),
            }

        claw_mat = Material.objects.filter(name='Claw Insulator', is_active=True).first()
        if claw_mat and claw_qty > 0:
            p = eff_price(claw_mat)
            material_label = claw_mat.name
            if num_hot_wires > 1:
                material_label = f"{claw_mat.name} (for {num_hot_wires} hot wires)"
            elif num_hot_wires == 1:
                material_label = f"{claw_mat.name} (for hot wire)"
            material_costs['insulators_claw'] = {
                'material': material_label,
                'unit_price': float(quantize_2(p)),
                'quantity': claw_qty,
                'cost': float(quantize_2(p * Decimal(claw_qty))),
            }

    # Strainers - 1 per 100m of fence length (same as stay posts)
    interval_m = Decimal('50') if netting_type == 'deer' else Decimal('100')
    total_recommended_strainers = int((fence_length / interval_m).to_integral_value(rounding=ROUND_UP))
    if fence_length > 0:
        total_recommended_strainers = max(total_recommended_strainers, 2)
    # For backward compatibility, calculate intermediate strainers
    additional_strainers = total_recommended_strainers - 2
    if additional_strainers < 0:
        additional_strainers = 0

    recommended_strainers_info = {
        'terrain': 'standard',
        'interval_meters_used': float(interval_m),
        'recommended_strainers_total': total_recommended_strainers,
        'recommended_strainers_intermediate': additional_strainers,
    }

    if fence_type.requires_strainers:
        strainer_mat = Material.objects.filter(name__iexact='2.5/7 inch Strainer', is_active=True).first()
        if strainer_mat:
            p = eff_price(strainer_mat)
            qty = total_recommended_strainers
            material_costs['strainers'] = {
                'material': strainer_mat.name,
                'unit_price': float(quantize_2(p)),
                'quantity': qty,
                'cost': float(quantize_2(p * Decimal(qty))),
            }

    # Stay posts (strainers) - 1 per 100m of fence length
    stay_posts_mat = Material.objects.filter(name__iexact='5 inch stay posts', is_active=True).first()
    if netting_type == 'deer':
        stay_posts_mat = Material.objects.filter(name__iexact='Deer Posts', is_active=True).first() or stay_posts_mat
    stay_interval = Decimal('50') if netting_type == 'deer' else Decimal('100')
    if stay_posts_mat and fence_length > 0:
        stay_posts_qty = int((fence_length / stay_interval).to_integral_value(rounding=ROUND_UP))
        stay_posts_qty = max(stay_posts_qty, 2)
        p = eff_price(stay_posts_mat)
        material_costs['stay_posts'] = {
            'material': stay_posts_mat.name,
            'unit_price': float(quantize_2(p)),
            'quantity': stay_posts_qty,
            'cost': float(quantize_2(p * Decimal(stay_posts_qty))),
        }

    # Triplex - 1 per 500m of fence length
    triplex_mat = Material.objects.filter(name='Triplex', is_active=True).first()
    if triplex_mat and fence_length > 0:
        triplex_qty = int((fence_length / Decimal('500')).to_integral_value(rounding=ROUND_UP))
        p = eff_price(triplex_mat)
        material_costs['triplex'] = {
            'material': triplex_mat.name,
            'unit_price': float(quantize_2(p)),
            'quantity': triplex_qty,
            'cost': float(quantize_2(p * Decimal(triplex_qty))),
        }

    total_material_cost = Decimal('0.00')
    for v in material_costs.values():
        total_material_cost += Decimal(str(v['cost']))
    total_material_cost = quantize_2(total_material_cost)

    total_cost = quantize_2(total_material_cost + labor_cost)

    # Optional electric outrigger components
    outrigger_details: Dict[str, Any] = {}
    if electric_outrigger:
        outrigger_wire_mat = Material.objects.filter(name__iexact='Electric Outrigger Wire', is_active=True).first()
        if not outrigger_wire_mat:
            outrigger_wire_mat = fence_type.wire_material

        if outrigger_wire_mat:
            roll_len = get_roll_length(outrigger_wire_mat)
            outrigger_rolls = Decimal(ceil_div(fence_length, roll_len))
            p = eff_price(outrigger_wire_mat)
            material_costs['outrigger_wire'] = {
                'material': outrigger_wire_mat.name,
                'unit_price': float(quantize_2(p)),
                'quantity': float(outrigger_rolls),
                'cost': float(quantize_2(p * outrigger_rolls)),
            }
            outrigger_details['wire_rolls'] = float(outrigger_rolls)

        insulator_mat = Material.objects.filter(name__iexact='Outrigger Insulator', is_active=True).first()
        if insulator_mat:
            insulator_qty = posts_required
            p = eff_price(insulator_mat)
            material_costs['outrigger_insulators'] = {
                'material': insulator_mat.name,
                'unit_price': float(quantize_2(p)),
                'quantity': insulator_qty,
                'cost': float(quantize_2(p * Decimal(insulator_qty))),
            }
            outrigger_details['insulators'] = insulator_qty

        connector_mat = Material.objects.filter(name__iexact='Outrigger Connectors', is_active=True).first()
        if connector_mat:
            connector_qty = max(2, int(posts_required / 50) + 1)
            p = eff_price(connector_mat)
            material_costs['outrigger_connectors'] = {
                'material': connector_mat.name,
                'unit_price': float(quantize_2(p)),
                'quantity': connector_qty,
                'cost': float(quantize_2(p * Decimal(connector_qty))),
            }
            outrigger_details['connectors'] = connector_qty

    return {
        'posts_required': posts_required,
        'post_spacing_used': float(quantize_2(Decimal(spacing))),
        'wire_count_used': total_wires,
        'wire_length_meters': float(wire_length_meters),
        'wire_rolls_required': float(wire_rolls_required),
        'labor_hours': float(labor_hours),
        'labor_rate_per_hour': float(lr),
        'build_rate_per_hour': float(br),
        'labor_cost': float(labor_cost),
        'material_costs': material_costs,
        'total_material_cost': float(total_material_cost),
        'total_cost': float(total_cost),
        'insulator_counts': insulator_counts,
        'strainer_recommendation': recommended_strainers_info,
        'staple_counts': staple_counts,
        'netting_type': netting_type,
        'netting_height_cm': float(netting_height_cm) if netting_height_cm else None,
        'netting_rolls_required': float(netting_rolls_required),
        'electric_outrigger': electric_outrigger,
        'electric_outrigger_details': outrigger_details,
    }


# -------- Reports (PDF / Excel) -------- #
from io import BytesIO
from reportlab.lib.pagesizes import A4
from reportlab.lib.units import mm
from reportlab.pdfgen import canvas
from reportlab.lib import colors
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, numbers


def combine_duplicate_materials(material_costs: dict) -> dict:
    """
    Combine materials with the same name in material_costs.
    This handles cases where wire materials appear twice (e.g., standard + hot wire using same material).
    """
    combined: dict = {}
    # Map of normalized material name -> key stored in `combined`
    index_by_norm: dict = {}

    def norm(name: str) -> str:
        # Normalize for grouping (ignore case and surrounding whitespace)
        return (name or '').strip().lower()

    for key, item in material_costs.items():
        material_name = item.get('material', '')
        normalized = norm(material_name)
        unit_price = item.get('unit_price', 0)
        quantity = item.get('quantity', 0)
        cost = item.get('cost', 0)

        if normalized in index_by_norm:
            # Already have this material grouped, add quantities and costs
            existing_key = index_by_norm[normalized]
            existing_item = combined[existing_key]
            existing_item['quantity'] = existing_item.get('quantity', 0) + quantity
            existing_item['cost'] = existing_item.get('cost', 0) + cost
            # Keep the first unit_price encountered to avoid small float diffs creating noise
        else:
            # New material group
            combined[key] = {
                'material': material_name.strip() if isinstance(material_name, str) else material_name,
                'unit_price': unit_price,
                'quantity': quantity,
                'cost': cost,
            }
            index_by_norm[normalized] = key

    return combined


def generate_pdf(calculation: FenceCalculation) -> bytes:
    buffer = BytesIO()
    c = canvas.Canvas(buffer, pagesize=A4)
    width, height = A4

    # Header
    c.setFillColor(colors.black)
    c.setFont("Helvetica-Bold", 16)
    c.drawString(30 * mm, height - 30 * mm, "Farm Fence Planner - Calculation Report")
    c.setFont("Helvetica", 10)
    c.drawString(30 * mm, height - 36 * mm, f"Fence Length: {calculation.fence_length} m")
    c.drawString(30 * mm, height - 42 * mm, f"Created: {timezone.localtime(calculation.created_at).strftime('%Y-%m-%d %H:%M')}")

    meta_y = height - 50 * mm
    c.setFont("Helvetica", 10)
    if calculation.netting_type and calculation.netting_type != 'none':
        netting_label = 'Deer' if calculation.netting_type == 'deer' else 'Sheep'
        c.drawString(30 * mm, meta_y, f"Netting Type: {netting_label}")
        meta_y -= 6 * mm
        if calculation.netting_height_cm:
            c.drawString(30 * mm, meta_y, f"Netting Height: {calculation.netting_height_cm} cm")
            meta_y -= 6 * mm
    if calculation.electric_outrigger:
        c.drawString(30 * mm, meta_y, "Electric Outrigger: Yes")
        meta_y -= 6 * mm
    else:
        c.drawString(30 * mm, meta_y, "Electric Outrigger: No")
        meta_y -= 6 * mm

    # Materials table
    y = meta_y - 6 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawString(30 * mm, y, "Materials (NZD excl. GST)")
    y -= 6 * mm
    c.setFont("Helvetica", 10)

    headers = ["Item", "Unit Price", "Qty", "Cost"]
    cols_x = [30 * mm, 100 * mm, 130 * mm, 160 * mm]
    for i, h in enumerate(headers):
        c.drawString(cols_x[i], y, h)
    y -= 5 * mm
    c.line(30 * mm, y, 190 * mm, y)
    y -= 3 * mm

    # Combine duplicate materials before displaying
    combined_materials = combine_duplicate_materials(calculation.material_costs)
    
    for k, v in combined_materials.items():
        c.drawString(cols_x[0], y, str(v.get('material', k)))
        c.drawRightString(cols_x[1] + 20 * mm, y, f"${v.get('unit_price', 0):.2f}")
        c.drawRightString(cols_x[2] + 10 * mm, y, f"{v.get('quantity', 0)}")
        c.drawRightString(cols_x[3] + 20 * mm, y, f"${v.get('cost', 0):.2f}")
        y -= 6 * mm
        if y < 30 * mm:
            c.showPage()
            y = height - 30 * mm

    # Totals
    y -= 4 * mm
    c.setFont("Helvetica-Bold", 12)
    c.drawRightString(160 * mm, y, "Material Total:")
    c.drawRightString(190 * mm, y, f"${calculation.total_material_cost:.2f}")
    y -= 6 * mm
    c.drawRightString(160 * mm, y, "Labor Cost:")
    c.drawRightString(190 * mm, y, f"${calculation.labor_cost:.2f}")
    y -= 6 * mm
    c.drawRightString(160 * mm, y, "Total (excl. GST):")
    c.drawRightString(190 * mm, y, f"${calculation.total_cost:.2f}")

    c.showPage()
    c.save()
    pdf = buffer.getvalue()
    buffer.close()
    return pdf


def generate_excel(calculation: FenceCalculation) -> bytes:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Fence Calculation"

    ws["A1"] = "Farm Fence Planner - Calculation Report"
    ws["A1"].font = Font(size=16, bold=True)
    ws.append(["Fence Length (m)", float(calculation.fence_length)])
    ws.append(["Created", timezone.localtime(calculation.created_at).strftime('%Y-%m-%d %H:%M')])

    if calculation.netting_type and calculation.netting_type != 'none':
        netting_label = 'Deer' if calculation.netting_type == 'deer' else 'Sheep'
        ws.append(["Netting Type", netting_label])
        if calculation.netting_height_cm:
            ws.append(["Netting Height (cm)", float(calculation.netting_height_cm)])
    else:
        ws.append(["Netting Type", "None"])

    ws.append(["Electric Outrigger", "Yes" if calculation.electric_outrigger else "No"])
    ws.append([])

    ws.append(["Item", "Unit Price (NZD)", "Qty", "Cost (NZD)"])
    header_fill = PatternFill("solid", fgColor="DDDDDD")
    for cell in ws[5]:
        cell.font = Font(bold=True)
        cell.fill = header_fill

    currency_fmt = numbers.FORMAT_CURRENCY_USD_SIMPLE.replace("$", "$")

    # Combine duplicate materials before displaying
    combined_materials = combine_duplicate_materials(calculation.material_costs)
    
    for k, v in combined_materials.items():
        ws.append([
            str(v.get('material', k)),
            float(v.get('unit_price', 0.0)),
            v.get('quantity', 0),
            float(v.get('cost', 0.0)),
        ])

    ws.append([])
    ws.append(["Material Total", None, None, float(calculation.total_material_cost)])
    ws.append(["Labor Cost", None, None, float(calculation.labor_cost)])
    ws.append(["Total (excl. GST)", None, None, float(calculation.total_cost)])

    # Format currency columns
    for row in ws.iter_rows(min_row=6, min_col=2, max_col=2):
        for cell in row:
            cell.number_format = currency_fmt
    for row in ws.iter_rows(min_row=6, min_col=4, max_col=4):
        for cell in row:
            cell.number_format = currency_fmt

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()
